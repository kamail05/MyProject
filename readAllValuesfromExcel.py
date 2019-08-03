# Importing openpyl package -> Creating openpyxl obj to open workbook -> creating sheet obj to access sheets
# -> Using sheet obj & excel cell value index/creating obj for cell and accessing value using its index
# -> To Read all the values present in excel -> using max_row & max_column to identify the no. of columns/rows
# which has values in the excel

import openpyxl

# Creating object for opening the excel file
wb = openpyxl.load_workbook('C:/Users/Abilash/PycharmProjects/AT/Testdata.xlsx')

# Creating obj for sheet to use the sheet we want
sh = wb['Hi Dude']

# To read data from all the rows and all the columns by finding how many rows/columns has values
rows = sh.max_row
columns = sh.max_column

print('Total No. of rows are',rows)
print('Total No. of Columns are',columns)

# Creating loop to fetch all the values from the excel
# Here we have given rows+1 because loop will skip the last value to avoid those issue +1 is added.
# After getting the row size & column size with the help of it we are iterating the for loop of rows & columns.
# Finally printing the cell values using object name.cell(row,column)(i,j)
for i in range(1,rows+1):
    for j in range(1,columns+1):
        c = sh.cell(i,j)
        print(c.value)

# Another way of achieving the getting all values from excel
# This method is for knowing the exact starting & ending cell value.
# Here we are directly iterating over the sheet by giving its start and end value of cell value index
# First iterating rows in excel and then with the iterated rows we are again iterating the column in it
# After iterating simply printing the values from column.value
for r in sh['A1':'E4']:
    for c in r:
        print(c.value)
# This method is used for given particular range
















